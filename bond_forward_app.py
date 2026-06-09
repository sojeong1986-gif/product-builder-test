import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook

st.set_page_config(
    page_title="동업사 본드포워드 현황",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
.stApp { background-color: #ffffff; }
[data-testid="stAppViewContainer"] { background-color: #ffffff; }
[data-testid="stHeader"] { background-color: #ffffff; }
.page-title { font-size: 22px; font-weight: 700; color: #1a2340; margin: 0; }
.page-subtitle { font-size: 13px; color: #8a94a6; margin-top: 4px; }
.kpi-card {
    background: #f7f9fc; border: 1px solid #e8edf3;
    border-radius: 12px; padding: 20px 22px;
}
.kpi-label { font-size: 12px; color: #8a94a6; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px; }
.kpi-value { font-size: 26px; font-weight: 700; color: #1a2340; line-height: 1.1; }
.kpi-sub { font-size: 12px; color: #8a94a6; margin-top: 4px; }
.kpi-badge { display: inline-block; padding: 2px 8px; border-radius: 4px;
    font-size: 11px; font-weight: 600; margin-top: 6px; }
.badge-up { background: #e8f5e9; color: #2e7d32; }
.badge-down { background: #ffebee; color: #c62828; }
.section-title { font-size: 15px; font-weight: 700; color: #1a2340;
    margin-bottom: 14px; padding-bottom: 8px; border-bottom: 1px solid #e8edf3; }
.styled-table { width: 100%; border-collapse: collapse; font-size: 13px; }
.styled-table thead tr { background: #f0f4f9; }
.styled-table th { padding: 10px 14px; text-align: right; font-weight: 600;
    color: #4a5568; border-bottom: 1px solid #e8edf3; white-space: nowrap; }
.styled-table th:first-child { text-align: left; }
.styled-table td { padding: 9px 14px; text-align: right; color: #2d3748;
    border-bottom: 1px solid #f0f4f9; white-space: nowrap; }
.styled-table td:first-child { text-align: left; font-weight: 600; color: #1a2340; }
.styled-table tr:hover { background: #f7f9fc; }
.styled-table tr.group-total { background: #e8f0fb; font-weight: 700; }
.styled-table tr.grand-total { background: #1a2340; color: white; font-weight: 700; }
.styled-table tr.grand-total td { color: white; border-bottom: none; }
@media print {
    .no-print { display: none !important; }
    .stApp { background: white; }
}
</style>
""", unsafe_allow_html=True)

# ── 엑셀 파싱 함수 ──────────────────────────────────────────────────
@st.cache_data
def load_excel(path="동업사_본드포워드_현황.xlsx"):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = [row for row in ws.iter_rows(values_only=True)]
    wb.close()

    # 기간 헤더 추출 (3행: index 2)
    header_row = data[2]
    periods = []
    for cell in header_row:
        if cell and ("년" in str(cell)) and ("월" in str(cell)):
            p = str(cell).strip()
            if p not in periods:
                periods.append(p)

    # 제외할 행 키워드
    skip_keywords = ["합계", "생/손보", "구분", "보험사", "본드포워드", "개별", "자산대비", None, ""]

    손보 = ["메리츠화재","삼성화재","DB손해보험","현대해상","KB손해보험",
             "한화손해보험","롯데손해보험","흥국화재","NH농협손해보험"]
    생보_names = ["삼성생명","한화생명","교보생명","신한라이프","농협생명",
                  "DB생명","미래에셋생명","동양생명","흥국생명"]

    companies = {}
    for row in data:
        name = str(row[0]).strip() if row[0] else ""
        if not name or any(kw in name for kw in skip_keywords):
            continue
        if name not in 손보 and name not in 생보_names:
            continue

        # 자산총계
        asset = None
        for cell in row[1:3]:
            try:
                v = float(str(cell).replace(",",""))
                if v > 1000:
                    asset = v
                    break
            except:
                pass

        # 잔액 (숫자 컬럼에서 기간 수만큼 추출)
        nums = []
        for cell in row[1:]:
            try:
                v = float(str(cell).replace(",",""))
                if 100 < v < 1000000:
                    nums.append(v)
            except:
                pass

        # 잔액은 앞에서 len(periods)개
        balances = nums[:len(periods)] if len(nums) >= len(periods) else nums

        구분 = "손보" if name in 손보 else "생보"
        companies[name] = {
            "구분": 구분,
            "자산총계": asset or 0,
            "잔액": balances
        }

    return companies, periods

# ── 데이터 로드 ─────────────────────────────────────────────────────
try:
    RAW, PERIODS = load_excel()
    if not RAW:
        st.error("❌ 엑셀에서 데이터를 읽지 못했습니다. 파일 형식을 확인해주세요.")
        st.stop()
except FileNotFoundError:
    st.error("❌ `동업사_본드포워드_현황.xlsx` 파일을 찾을 수 없습니다. GitHub 레포에 파일이 있는지 확인해주세요.")
    st.stop()

def make_df():
    rows = []
    for co, d in RAW.items():
        for i, p in enumerate(PERIODS):
            if i >= len(d["잔액"]):
                continue
            bal = d["잔액"][i]
            rows.append({
                "회사": co, "구분": d["구분"],
                "자산총계": d["자산총계"],
                "기간": p, "잔액": bal,
                "비중": round(bal / d["자산총계"] * 100, 2) if d["자산총계"] else 0
            })
    return pd.DataFrame(rows)

df_long = make_df()

# ── 헤더 ───────────────────────────────────────────────────────────
col_title, col_btn = st.columns([6, 1])
with col_title:
    st.markdown("""
    <div>
        <div class="page-title">📊 동업사 본드포워드 현황</div>
        <div class="page-subtitle">보험업계 금리선도(Bond Forward) 미결제약정 현황 모니터링</div>
    </div>
    """, unsafe_allow_html=True)
with col_btn:
    st.markdown("""
    <div style="display:flex;justify-content:flex-end;align-items:center;height:100%;padding-top:12px;">
        <button class="no-print" onclick="window.print()"
            style="padding:8px 18px;background:#1a2340;color:white;border:none;
                   border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">
            🖨️ 인쇄 / PDF
        </button>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# ── 필터 ───────────────────────────────────────────────────────────
fc1, fc2, fc3 = st.columns([2, 2, 4])
with fc1:
    sel_period = st.selectbox("📅 조회 기준 시점", PERIODS, index=len(PERIODS) - 1)
with fc2:
    sel_type = st.selectbox("🏢 구분", ["전체", "손보", "생보"])
with fc3:
    sel_companies = st.multiselect("🔍 회사 선택 (미선택 시 전체)", options=list(RAW.keys()), default=[])

# 필터 적용
df_cur = df_long[df_long["기간"] == sel_period].copy()
손보_bal = df_long[(df_long["기간"] == sel_period) & (df_long["구분"] == "손보")]["잔액"].sum()
생보_bal = df_long[(df_long["기간"] == sel_period) & (df_long["구분"] == "생보")]["잔액"].sum()
total_bal = 손보_bal + 생보_bal

st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

# ── KPI 카드 ────────────────────────────────────────────────────────
prev_idx = PERIODS.index(sel_period) - 1
if prev_idx >= 0:
    prev_period = PERIODS[prev_idx]
    prev_bal = df_long[df_long["기간"] == prev_period]["잔액"].sum()
    chg = total_bal - prev_bal
    chg_pct = chg / prev_bal * 100 if prev_bal else 0
    chg_html = f'<span class="kpi-badge {"badge-up" if chg>=0 else "badge-down"}">{"▲" if chg>=0 else "▼"} {abs(chg_pct):.1f}%</span>'
else:
    chg_html = ""

top_co = df_long[df_long["기간"] == sel_period].sort_values("비중", ascending=False).iloc[0]

k1, k2, k3, k4 = st.columns(4)
with k1:
    st.markdown(f"""<div class="kpi-card">
        <div class="kpi-label">생/손보 합계</div>
        <div class="kpi-value">{total_bal/1000:,.1f}조</div>
        <div class="kpi-sub">{total_bal:,.0f} 억원</div>{chg_html}
    </div>""", unsafe_allow_html=True)
with k2:
    st.markdown(f"""<div class="kpi-card">
        <div class="kpi-label">손보 합계</div>
        <div class="kpi-value">{손보_bal/1000:,.1f}조</div>
        <div class="kpi-sub">{손보_bal:,.0f} 억원 · {손보_bal/total_bal*100:.1f}%</div>
    </div>""", unsafe_allow_html=True)
with k3:
    st.markdown(f"""<div class="kpi-card">
        <div class="kpi-label">생보 합계</div>
        <div class="kpi-value">{생보_bal/1000:,.1f}조</div>
        <div class="kpi-sub">{생보_bal:,.0f} 억원 · {생보_bal/total_bal*100:.1f}%</div>
    </div>""", unsafe_allow_html=True)
with k4:
    st.markdown(f"""<div class="kpi-card">
        <div class="kpi-label">자산대비 비중 1위</div>
        <div class="kpi-value" style="font-size:20px">{top_co['회사']}</div>
        <div class="kpi-sub">자산대비 {top_co['비중']:.2f}%</div>
    </div>""", unsafe_allow_html=True)

st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

# ── 차트 + 테이블 ──────────────────────────────────────────────────
chart_col, table_col = st.columns([5, 5])

with chart_col:
    st.markdown('<div class="section-title">📈 본드포워드 잔액 추이</div>', unsafe_allow_html=True)

    손보_trend = df_long[df_long["구분"] == "손보"].groupby("기간")["잔액"].sum().reindex(PERIODS)
    생보_trend = df_long[df_long["구분"] == "생보"].groupby("기간")["잔액"].sum().reindex(PERIODS)

    fig = go.Figure()
    if sel_companies:
        colors = px.colors.qualitative.Plotly
        for i, co in enumerate(sel_companies):
            vals = [df_long[(df_long["회사"] == co) & (df_long["기간"] == p)]["잔액"].values[0]
                    if len(df_long[(df_long["회사"] == co) & (df_long["기간"] == p)]) > 0 else 0
                    for p in PERIODS]
            fig.add_trace(go.Scatter(x=PERIODS, y=vals, name=co, mode="lines+markers",
                line=dict(color=colors[i % len(colors)], width=2), marker=dict(size=6)))
    else:
        fig.add_trace(go.Bar(x=PERIODS, y=손보_trend.values, name="손보",
            marker_color="#1565c0", opacity=0.85))
        fig.add_trace(go.Bar(x=PERIODS, y=생보_trend.values, name="생보",
            marker_color="#6a1b9a", opacity=0.85))
        fig.add_trace(go.Scatter(x=PERIODS, y=(손보_trend + 생보_trend).values, name="합계",
            mode="lines+markers", line=dict(color="#e53935", width=2.5), marker=dict(size=7)))

    fig.update_layout(
        barmode="stack", plot_bgcolor="white", paper_bgcolor="white",
        font=dict(family="sans-serif", size=12, color="#2d3748"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    bgcolor="rgba(255,255,255,0.9)", bordercolor="#e8edf3", borderwidth=1),
        xaxis=dict(showgrid=False, tickfont=dict(size=11)),
        yaxis=dict(showgrid=True, gridcolor="#f0f4f9", tickformat=",",
                   ticksuffix=" 억", tickfont=dict(size=11)),
        margin=dict(l=10, r=10, t=40, b=10), height=340, hovermode="x unified",
    )
    fig.add_vline(x=sel_period, line_dash="dash", line_color="#ff7043", line_width=1.5,
                  annotation_text="조회시점", annotation_font_color="#ff7043", annotation_font_size=11)
    st.plotly_chart(fig, use_container_width=True)

    # 도넛
    st.markdown('<div class="section-title" style="margin-top:4px">🥧 생/손보 비중</div>', unsafe_allow_html=True)
    fig2 = go.Figure(go.Pie(
        labels=["손보", "생보"], values=[손보_bal, 생보_bal], hole=0.6,
        marker_colors=["#1565c0", "#6a1b9a"], textinfo="label+percent", textfont_size=13,
    ))
    fig2.update_layout(
        showlegend=False, plot_bgcolor="white", paper_bgcolor="white",
        margin=dict(l=10, r=10, t=10, b=10), height=210,
        annotations=[dict(text=f"{total_bal/10000:.1f}조", x=0.5, y=0.5,
                          font_size=18, font_color="#1a2340", showarrow=False)]
    )
    st.plotly_chart(fig2, use_container_width=True)

with table_col:
    st.markdown(f'<div class="section-title">📋 상세 현황 ({sel_period})</div>', unsafe_allow_html=True)

    def render_group(구분):
        cos = [c for c, d in RAW.items() if d["구분"] == 구분]
        if sel_companies:
            cos = [c for c in cos if c in sel_companies]
        rows_html = ""
        grp_bal = 0
        for co in cos:
            d = RAW[co]
            pi = PERIODS.index(sel_period)
            cur = d["잔액"][pi] if pi < len(d["잔액"]) else 0
            prev = d["잔액"][pi - 1] if pi > 0 and (pi - 1) < len(d["잔액"]) else cur
            diff = cur - prev
            arrow = "▲" if diff > 0 else ("▼" if diff < 0 else "–")
            color = "#2e7d32" if diff > 0 else ("#c62828" if diff < 0 else "#888")
            ratio = cur / d["자산총계"] * 100 if d["자산총계"] else 0
            rows_html += f"""<tr>
                <td>{co}</td>
                <td>{d['자산총계']:,.0f}</td>
                <td>{cur:,.0f}</td>
                <td><span style='color:{color};font-size:11px'>{arrow} {abs(diff):,.0f}</span></td>
                <td>{ratio:.2f}%</td>
            </tr>"""
            grp_bal += cur
        share = grp_bal / total_bal * 100 if total_bal else 0
        rows_html += f"""<tr class="group-total">
            <td>{'손보' if 구분=='손보' else '생보'} 합계</td>
            <td></td><td>{grp_bal:,.0f}</td><td></td><td>{share:.1f}%</td>
        </tr>"""
        return rows_html

    bg1, fg1 = "#e3f2fd", "#1565c0"
    bg2, fg2 = "#f3e5f5", "#6a1b9a"
    손보_rows = f'<tr><td colspan="5" style="background:{bg1};color:{fg1};font-weight:700;padding:6px 14px;font-size:12px">▶ 손해보험</td></tr>' + render_group("손보")
    생보_rows = f'<tr><td colspan="5" style="background:{bg2};color:{fg2};font-weight:700;padding:6px 14px;font-size:12px">▶ 생명보험</td></tr>' + render_group("생보")
    grand = f'<tr class="grand-total"><td>생/손보 합계</td><td></td><td>{total_bal:,.0f}</td><td></td><td>100%</td></tr>'

    header = """<table class="styled-table"><thead><tr>
        <th style="text-align:left">회사</th><th>자산총계(억)</th>
        <th>잔액(억)</th><th>전분기比</th><th>자산비중</th>
    </tr></thead><tbody>"""

    body = {"전체": 손보_rows + 생보_rows + grand, "손보": 손보_rows, "생보": 생보_rows}
    st.markdown(header + body[sel_type] + "</tbody></table>", unsafe_allow_html=True)

# ── 기간별 전체 추이 테이블 ──────────────────────────────────────────
st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
st.markdown('<div class="section-title">📅 기간별 잔액 추이 (억원)</div>', unsafe_allow_html=True)

cos_show = list(RAW.keys())
if sel_companies:
    cos_show = sel_companies
elif sel_type != "전체":
    cos_show = [c for c, d in RAW.items() if d["구분"] == sel_type]

pivot_rows = ""
for 구분, bg, fg, label in [("손보", "#e3f2fd", "#1565c0", "손해보험"), ("생보", "#f3e5f5", "#6a1b9a", "생명보험")]:
    cos_g = [c for c in cos_show if RAW[c]["구분"] == 구분]
    if not cos_g:
        continue
    pivot_rows += f'<tr><td colspan="{len(PERIODS)+2}" style="background:{bg};color:{fg};font-weight:700;padding:6px 14px;font-size:12px">▶ {label}</td></tr>'
    grp_sums = [0] * len(PERIODS)
    for co in cos_g:
        d = RAW[co]
        cells = ""
        for i, p in enumerate(PERIODS):
            bal = d["잔액"][i] if i < len(d["잔액"]) else 0
            grp_sums[i] += bal
            hl = "background:#fff8e1;font-weight:700;" if p == sel_period else ""
            cells += f'<td style="{hl}">{bal:,.0f}</td>'
        pivot_rows += f'<tr><td>{co}</td><td>{d["자산총계"]:,.0f}</td>{cells}</tr>'
    gcells = "".join([f'<td style="{"background:#fff8e1;" if p==sel_period else ""}">{s:,.0f}</td>'
                      for p, s in zip(PERIODS, grp_sums)])
    pivot_rows += f'<tr class="group-total"><td>{"손보" if 구분=="손보" else "생보"} 소계</td><td></td>{gcells}</tr>'

ph = "<th style='text-align:left'>회사</th><th>자산총계</th>" + "".join([
    f'<th style="{"color:#e65100;font-weight:800;" if p==sel_period else ""}">{p}</th>' for p in PERIODS])
st.markdown(f'<table class="styled-table"><thead><tr>{ph}</tr></thead><tbody>{pivot_rows}</tbody></table>',
            unsafe_allow_html=True)

# ── 하단 ────────────────────────────────────────────────────────────
st.markdown("""
<div style="margin-top:32px;padding-top:16px;border-top:1px solid #e8edf3;
     display:flex;justify-content:space-between;align-items:center;">
    <span style="font-size:12px;color:#aab0bc;">
        ※ 자료: 각사 DART 공시 / 단위: 억원, % &nbsp; ※ 자산대비 비중 = 본드포워드 잔액 ÷ 자산총계
    </span>
    <span style="font-size:12px;color:#aab0bc;">메리츠화재 자산운용본부</span>
</div>
""", unsafe_allow_html=True)
